from openpyxl import load_workbook
import config as cfg
from textProcessor import TextProcessor
from interactable import Button
from gameNode import GameNode, CellData
import pygame as pg

class Game:
    def __init__(self, gamefile):
        self.file_wb = load_workbook(gamefile)
        self.file_text_ws = self.file_wb['text']

        # building every node in a row and then hooking them up to eachother
        self.options = self.build_nodes()
        

    def build_nodes(self):
        end_nodes = []
        options = {}
        cell = self.file_text_ws.cell
        for i in range(2,self.file_text_ws.max_row):
            # acquiring option data
            option_node = GameNode(row_id=i,option_id=cell(i, 1).value,data=CellData(conditional=cell(i, 3).value,destination=cell(i, 4).value,incoming=cell(i, 17).value,leaving=cell(i, 18).value,ending=cell(i, 19).value,text = self.file_text_ws.cell(i, 2).value, noise = self.file_wb['sounds'].cell(i, 2).value, models = self.file_wb['models'].cell(i, 2).value, bg = self.file_wb['bgs'].cell(i, 2).value, music = self.file_wb['music'].cell(i, 2).value))
            if str(cell(i,1).value) not in options:
                options[str(cell(i,1).value)] = [option_node]
            else:
                options[str(cell(i,1).value)].append(option_node) 
            prev_option = option_node

            # acquiring story text data. Regular story nodes will point to the next story node, so they can point to the next node
            for j in range(5,17):
                if self.file_text_ws.cell(i, j).value != None:
                    text_node = GameNode(row_id=i,data=CellData(conditional=cell(i, 3).value, destination=cell(i, 4).value, incoming=cell(i, 17).value, leaving=cell(i, 18).value, ending=cell(i, 19).value, text = self.file_text_ws.cell(i, j).value, noise = self.file_wb['sounds'].cell(i, j).value, models = self.file_wb['models'].cell(i, j).value, bg = self.file_wb['bgs'].cell(i, j).value, music = self.file_wb['music'].cell(i, j).value))
                    prev_option.add_child(text_node)
                    prev_option = text_node
                if self.file_text_ws.cell(i, j).value == None or j == 17:
                    prev_option.last_text_box_bool = True
                    end_nodes.append(prev_option)
                    break

        keys = {0:pg.K_1, 1:pg.K_2, 2:pg.K_3, 3:pg.K_4,}

        # building buttons for options dicitionary
        for key, value_list in options.items():
            if len(value_list) == 1:
                value_list = value_list[0].get_children()
            else:
                BUTTON_WIDTH = round(cfg.WINDOW_WIDTH / (len(value_list) + 1))
                BUTTON_MARGIN = round((cfg.WINDOW_WIDTH - BUTTON_WIDTH * len(value_list)) / (len(value_list) + 1))
                for index, option in enumerate(value_list):
                    if option.data.text == None:
                        option = option.get_children()[0]
                    else:
                        option.box_dimension = (BUTTON_WIDTH * index + BUTTON_MARGIN * (index + 1), cfg.BUTTON_Y, BUTTON_WIDTH, cfg.BUTTON_BASE_HEIGHT)
                        option.button_dimension = BUTTON_WIDTH * 0.9, cfg.BUTTON_BASE_HEIGHT * 0.85, BUTTON_WIDTH * 0.72, cfg.BUTTON_BASE_HEIGHT * 0.425
                        option.trigger = [pg.MOUSEBUTTONDOWN, keys[index]]
                    value_list[index] = option
            options[key] = value_list
            
        return options